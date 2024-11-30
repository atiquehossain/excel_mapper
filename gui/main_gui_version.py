import os
import re
from tkinter import filedialog, messagebox, Tk, Label, Button, Frame,BooleanVar, Checkbutton
from tkinter.ttk import Combobox
from openpyxl import load_workbook
import pandas as pd
from num2words import num2words



# Utility Class
class Utils:
    """Utility methods for processing strings and data."""

    @staticmethod
    def convert_number_to_text(value):
        """Convert numeric values to text."""
        try:
            if isinstance(value, (int, float)):
                return num2words(value)
            return value
        except Exception as e:
            messagebox.showerror("Error",f"Error converting {value}: {e}")
            return value

    @staticmethod
    def clean_column_name(name):
        """Clean column names by replacing non-alphanumeric characters with underscores."""
        return re.sub(r'[^0-9a-zA-Z]+', '_', str(name).strip()).lower()

    @staticmethod
    def sanitize_key(name):
        """Sanitize keys to ensure they are valid Dart identifiers."""
        return re.sub(r'[^0-9a-zA-Z]+', '_', str(name).strip()).lower()

    @staticmethod
    def write_dart_file(folder, file_name, content):
        """Write Dart code to a file."""
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, file_name)
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content)


# Excel Processor
class ExcelProcessor:
    """Handles processing of the Excel file and data extraction."""

    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.df = None

    def load_sheet(self):
        """Loads and cleans the Excel sheet."""
        workbook = load_workbook(self.file_path, data_only=True)
        sheet = workbook[self.sheet_name]

        # Identify hidden rows
        hidden_rows = [row for row in sheet.row_dimensions if sheet.row_dimensions[row].hidden]

        # Load data with pandas and drop hidden rows
        self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
        self.df = self.df.drop(index=[row - 1 for row in hidden_rows if row <= len(self.df)])

        # Clean column names
        self.df.columns = [
            Utils.clean_column_name(col) for col in self.df.columns
        ]

    def get_dataframe(self):
        """Returns the processed DataFrame."""
        return self.df


class DartCodeGenerator:
    """Generates Dart code for widgets, models, localization, key files, and field names."""

    FIELD_TYPE_MAPPINGS = {
        "dropdown": "AppConstant.FieldType_dropdown",
        "multiple_choice": "AppConstant.FieldType_multiple_choice",
        "yes/no": "AppConstant.FieldType_radio",
        "number": "AppConstant.FieldType_EditText",
        "text": "AppConstant.FieldType_EditText",
        "image": "AppConstant.FieldType_Image",
    }

    def __init__(self, class_name, df, output_folder='dart_output'):
        self.class_name = class_name
        self.df = df
        self.output_folder = output_folder
        self.localization_data = {"English": {}, "Tamil": {}, "Sinhala": {}}
        self.localization_keys = []
        self.dart_widgets = []
        self.model_fields = []
        self.constructor_params = []
        self.from_json = []
        self.to_json = []
        self.to_string = []
        self.field_names = []  # List to store field names

        os.makedirs(self.output_folder, exist_ok=True)

    def process_row(self, row, row_number):
        """Processes a single row of data."""
        try:
            # Retrieve the row data with default values to avoid KeyErrors
            question = row.get('questions_in_english', 'Missing value')
            label = row.get('labels_in_english', 'Missing value')
            tamil_question_translation = row.get('questions_in_tamil', 'Missing value')
            sinhala_question_translation = row.get('questions_in_sinhala', 'Missing value')
            tamil_label_translation = row.get('labels_in_tamil', 'Missing value')
            sinhala_label_translation = row.get('labels_in_sinhala', 'Missing value')
            field_name_english = row.get('field_names_in_english', 'Missing value')
            field_name_sinhala = row.get('field_names_in_sinhala', 'Missing value')
            field_name_tamil = row.get('field_names_in_tamil', 'Missing value')
            data_type = row.get('data_type', None)
            model = row.get('database', None)

            if not model or pd.isna(model):
                return

            # Determine the field type
            field_type = self.get_field_type(data_type)

            # Sanitize and store localization data for question and label
            question_key, label_key = self.process_localization_data(
                question, label, tamil_question_translation, 
                sinhala_question_translation, tamil_label_translation, 
                sinhala_label_translation
            )

            # Process field name localization for each row independently
            self.process_field_name_localization(field_name_english, field_name_sinhala, field_name_tamil)

            # Generate widget code
            self.generate_widget_code(row_number, field_type, model, question_key, label_key)

            # Handle model data
            self.handle_model_data(model)

        except Exception as e:
            messagebox.showerror("Error", f"Error processing row {row_number}: {e}")
            return

    def handle_model_data(self, model):
        """Handles model data for generating Dart code for the model class."""
        field_name = Utils.sanitize_key(model)
        self.model_fields.append(f"final String {field_name};")
        self.constructor_params.append(f"required this.{field_name}")
        self.from_json.append(f"{field_name}: json['{field_name}'],")
        self.to_json.append(f"'{field_name}': {field_name},")
        self.to_string.append(f"{field_name}: {field_name}")

    def generate_files(self):
        """Generates Dart files."""
        # Generate Widget file
        widget_code = f"""
        class {self.class_name}UI {{
          Widget build(BuildContext context) {{
            return Column(
              children: [
                {"".join(self.dart_widgets)}
              ],
            );
          }}
        }}
        """
        Utils.write_dart_file(self.output_folder, f"{self.class_name.lower()}_widget.dart", widget_code)

        # Generate Model file
        model_code = f"""
        class {self.class_name}Model {{
          {''.join(self.model_fields)}

          {self.class_name}Model({{
            {''.join(self.constructor_params)}
          }});

          factory {self.class_name}Model.fromJson(Map<String, dynamic> json) {{
            return {self.class_name}Model(
              {''.join(self.from_json)}
            );
          }}

          Map<String, dynamic> toJson() {{
            return {{
              {''.join(self.to_json)}
            }};
          }}

          @override
          String toString() {{
            return "{self.class_name}({{ {''.join(self.to_string)} }})";
          }}
        }}
        """
        Utils.write_dart_file(self.output_folder, f"{self.class_name.lower()}_model.dart", model_code)

        # Generate localization files
        self.generate_localization_files()

        # Generate Key file and Field Names Dart file
        self.generate_key_file()
        self.generate_field_names()

    def get_field_type(self, data_type):
        """Returns the field type based on the data_type."""
        if isinstance(data_type, str):
            return self.FIELD_TYPE_MAPPINGS.get(data_type.lower(), "AppConstant.FieldType_EditText")
        return "AppConstant.FieldType_EditText"  # Default field type

    def process_localization_data(self, question, label, tamil_question, sinhala_question, tamil_label, sinhala_label):
        """Processes the localization data (translations for different languages)."""
        # Generate keys based on the input text
        question_key = Utils.sanitize_key(question)
        label_key = Utils.sanitize_key(label)

        # Add translations to localization dictionary
        self.localization_data["English"][question_key] = question
        self.localization_data["English"][label_key] = label

        self.localization_data["Tamil"][question_key] = tamil_question
        self.localization_data["Tamil"][label_key] = tamil_label

        self.localization_data["Sinhala"][question_key] = sinhala_question
        self.localization_data["Sinhala"][label_key] = sinhala_label

        # Add keys to list
        if question_key not in self.localization_keys:
            self.localization_keys.append(question_key)
        if label_key not in self.localization_keys:
            self.localization_keys.append(label_key)

        return question_key, label_key

    def process_field_name_localization(self, field_name_english, field_name_sinhala, field_name_tamil):
        """Processes field name localization for each row independently."""
        # Generate keys based on the input field names
        field_key_english = Utils.sanitize_key(field_name_english)
        field_key_sinhala = Utils.sanitize_key(field_name_sinhala)
        field_key_tamil = Utils.sanitize_key(field_name_tamil)

        # Add translations for field names directly to the list for each row
        self.localization_data["English"][field_key_english] = field_name_english
        self.localization_data["Tamil"][field_key_tamil] = field_name_tamil
        self.localization_data["Sinhala"][field_key_sinhala] = field_name_sinhala

        # Add field names to the list for the field names Dart file
        self.field_names.append(field_key_english)

        # Add field keys to localization keys list
        if field_key_english not in self.localization_keys:
            self.localization_keys.append(field_key_english)
        if field_key_sinhala not in self.localization_keys:
            self.localization_keys.append(field_key_sinhala)
        if field_key_tamil not in self.localization_keys:
            self.localization_keys.append(field_key_tamil)

    def generate_widget_code(self, row_number, field_type, model, question_key, label_key):
        """Generates Dart code for a widget based on the row data."""
        # Example of generating a widget code (adjust as necessary for your use case)
        widget_code = f"""
        class {model}Widget{row_number} extends StatelessWidget {{
          @override
          Widget build(BuildContext context) {{
            return TextFormField(
              decoration: InputDecoration(
                labelText: AppLocalization.of(context)!.get("{label_key}"),
                hintText: AppLocalization.of(context)!.get("{question_key}"),
              ),
              keyboardType: {field_type},
            );
          }}
        }}
        """
        self.dart_widgets.append(widget_code)

    def generate_localization_files(self):
        """Generates localization files for English, Tamil, and Sinhala."""
        for lang, translations in self.localization_data.items():
            localization_code = f"""
            class {lang} {{
              {"".join([f"static const String {key} = '{value}';\n" for key, value in translations.items()])}
            }}
            """
            file_name = f"{lang.lower()}_localization.dart"
            Utils.write_dart_file(self.output_folder, file_name, localization_code)

    def generate_key_file(self):
        """Generates a Dart file for localization keys."""
        key_code = f"""
        class LocalizationKeys {{
          {"".join([f"static const String {key} = '{key}';\n" for key in self.localization_keys])}
        }}
        """
        Utils.write_dart_file(self.output_folder, "localization_keys.dart", key_code)

    def generate_field_names(self):
        """Generates a Dart file for field names."""
        field_names_code = f"""
        class FieldNames {{
          {"\n  ".join([f"static const String {field_name} = '{field_name}';" for field_name in self.field_names])}
        }}
        """
        Utils.write_dart_file(self.output_folder, f"{self.class_name.lower()}_field_names.dart", field_names_code)

# GUI Class with Improved Look
class DartCodeGeneratorGUI:
    """GUI for Dart Code Generation."""

    def __init__(self, root):
        self.root = root
        self.root.title("Dart Code Generator")
        self.root.geometry("500x450")
        self.root.config(bg="#f5f5f5")

        # Create a Frame for better layout control
        self.frame = Frame(self.root, bg="#ffffff", padx=20, pady=20)
        self.frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Title Label
        self.title_label = Label(self.frame, text="Dart Code Generator", font=("Helvetica", 18, "bold"), bg="#ffffff", fg="#333333")
        self.title_label.grid(row=0, column=0, columnspan=2, pady=20)

        # Excel file selection
        self.excel_file_path = None
        self.sheet_name = None

        # File selection label and button
        self.file_label = Label(self.frame, text="No file selected", font=("Helvetica", 12), bg="#ffffff", fg="#777777")
        self.file_label.grid(row=1, column=0, columnspan=2, pady=10)

        self.select_file_button = Button(self.frame, text="Select Excel File", font=("Helvetica", 12), command=self.select_file, bg="#4CAF50", fg="white", relief="flat")
        self.select_file_button.grid(row=2, column=0, columnspan=2, pady=10, ipadx=10, ipady=5)

        # Sheet name dropdown
        self.sheet_label = Label(self.frame, text="Select Sheet Name", font=("Helvetica", 12), bg="#ffffff", fg="#333333")
        self.sheet_label.grid(row=3, column=0, pady=5)

        self.sheet_name_combobox = Combobox(self.frame, font=("Helvetica", 12), state="readonly")
        self.sheet_name_combobox.grid(row=3, column=1, pady=5, padx=10, ipadx=8)

        # Field names checkbox
        self.generate_field_names_var = BooleanVar(value=True)
        self.generate_field_names_checkbox = Checkbutton(self.frame, text="Generate Field Names", variable=self.generate_field_names_var, font=("Helvetica", 12), bg="#ffffff")
        self.generate_field_names_checkbox.grid(row=4, column=0, columnspan=2, pady=10)

        # Generate Dart files button
        self.generate_button = Button(self.frame, text="Generate Dart Files", font=("Helvetica", 12), command=self.generate_dart_files, bg="#008CBA", fg="white", relief="flat")
        self.generate_button.grid(row=5, column=0, columnspan=2, pady=20, ipadx=10, ipady=5)

    def select_file(self):
        """Allow the user to select an Excel file."""
        self.excel_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if self.excel_file_path:
            self.file_label.config(text=f"Selected File: {os.path.basename(self.excel_file_path)}")
            # Load the Excel file and populate the sheet dropdown
            self.load_sheets()

    def load_sheets(self):
        """Loads sheet names into the combobox."""
        try:
            workbook = load_workbook(self.excel_file_path)
            sheet_names = workbook.sheetnames
            self.sheet_name_combobox['values'] = sheet_names
            self.sheet_name_combobox.current(0)  # Set default to first sheet
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheets: {str(e)}")

    

    def generate_dart_files(self):
        """Trigger the Dart code generation."""
        self.sheet_name = self.sheet_name_combobox.get()
        if not self.excel_file_path or not self.sheet_name:
            messagebox.showerror("Error", "Please select a file and choose a sheet name.")
            return
        
        try:
            # Initialize Excel Processor
            excel_processor = ExcelProcessor(self.excel_file_path, self.sheet_name)
            excel_processor.load_sheet()
            df = excel_processor.get_dataframe()
            

            # Create Dart Code Generator with field names option
            class_name = self.sheet_name.replace(" ", "")
            generate_field_names = self.generate_field_names_var.get()
            generator = DartCodeGenerator(class_name, df)

            for index, row in df.iterrows():
                generator.process_row(row, index)

            # Generate the files based on user choice
            if generate_field_names:
                generator.generate_field_names()
            generator.generate_files()

            messagebox.showinfo("Success", "Dart files generated successfully!")

        except Exception as e:
            print( f"Error generating Dart files: {str(e)}")
            messagebox.showerror("Error", f"Error generating Dart files: {str(e)}")

# Start the application
if __name__ == "__main__":
    root = Tk()
    app = DartCodeGeneratorGUI(root)
    root.mainloop()
