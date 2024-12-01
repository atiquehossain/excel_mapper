import os
import re
from tkinter import Tk, Label, Button, Frame, filedialog, messagebox, Scrollbar, Listbox, MULTIPLE, StringVar
from tkinter.ttk import Combobox, Entry, Style, Button as TTKButton
from openpyxl import load_workbook
import pandas as pd


# Utility Class
class Utils:
    """Utility methods for processing strings and data."""

    @staticmethod
    def clean_column_name(name):
        """Clean column names by replacing non-alphanumeric characters with underscores."""
        return re.sub(r'[^0-9a-zA-Z]+', '_', str(name).strip()).lower()

    @staticmethod
    def sanitize_key(name):
        """Sanitize keys to ensure they are valid Dart identifiers."""
        return re.sub(r'[^0-9a-zAZ]+', '_', str(name).strip()).lower()

    @staticmethod
    def write_dart_file(folder, file_name, content):
        """Write Dart code to a file."""
        os.makedirs(folder, exist_ok=True)
        file_path = os.path.join(folder, file_name)
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content)


# GUI Class
class DartCodeGeneratorGUI:
    """GUI for Dart Code Generation."""

    def __init__(self, root):
        self.root = root
        self.root.title("Dynamic Dart Code Generator")
        self.root.geometry("900x800")
        self.root.config(bg="#f5f5f5")

        # Data Storage
        self.excel_file_path = None
        self.sheet_name = None
        self.df = None
        self.selected_columns = []
        self.column_mappings = {}
        self.languages = []
        self.key_language = None

        # Create GUI components
        self.create_gui()

    def create_gui(self):
        """Create GUI components."""
        style = Style()
        style.configure("TButton", font=("Helvetica", 12), padding=6)
        style.configure("TLabel", font=("Helvetica", 12), background="#f5f5f5")

        # File selection section
        Label(self.root, text="Step 1: Select Excel File", font=("Helvetica", 14, "bold"), bg="#f5f5f5").pack(pady=10)
        TTKButton(self.root, text="Select File", command=self.select_file, style="TButton").pack(pady=10)

        # File and sheet info
        self.file_label = Label(self.root, text="No file selected", bg="#f5f5f5", fg="#555555")
        self.file_label.pack(pady=5)

        self.sheet_label = Label(self.root, text="Select Sheet:", bg="#f5f5f5", fg="#333333")
        self.sheet_label.pack(pady=5)

        self.sheet_combobox = Combobox(self.root, state="readonly", width=30)
        self.sheet_combobox.pack(pady=5)
        self.sheet_combobox.bind("<<ComboboxSelected>>", self.on_sheet_selected)

        # Column selection and Key Language section
        Label(self.root, text="Step 2: Select Key Language and Map Columns", font=("Helvetica", 14, "bold"), bg="#f5f5f5").pack(pady=10)

        self.key_language_label = Label(self.root, text="Select Key Language:", bg="#f5f5f5", fg="#333333")
        self.key_language_label.pack(pady=5)

        self.key_language_combobox = Combobox(self.root, state="readonly", width=30)
        self.key_language_combobox.pack(pady=5)

        # Column selection section
        Label(self.root, text="Step 3: Select Columns to Map", font=("Helvetica", 14, "bold"), bg="#f5f5f5").pack(pady=10)

        frame = Frame(self.root)
        frame.pack(fill="both", expand=True, pady=10, padx=20)

        # Listbox for column selection
        self.column_listbox = Listbox(frame, selectmode=MULTIPLE, width=40, height=15)
        self.column_listbox.pack(side="left", fill="y")

        scrollbar = Scrollbar(frame, orient="vertical")
        scrollbar.config(command=self.column_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.column_listbox.config(yscrollcommand=scrollbar.set)

        # Additional Mapping Fields
        self.additional_fields_frame = Frame(self.root, bg="#f5f5f5")
        self.additional_fields_frame.pack(fill="both", expand=True, pady=10)

        # Buttons for mapping and generation
        TTKButton(self.root, text="Map Selected Columns", command=self.map_columns).pack(pady=10)
        TTKButton(self.root, text="Generate Dart Files", command=self.generate_dart_files).pack()

    def select_file(self):
        """Allow the user to select an Excel file."""
        self.excel_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if self.excel_file_path:
            self.file_label.config(text=f"Selected File: {os.path.basename(self.excel_file_path)}")
            self.load_sheets()

    def load_sheets(self):
        """Load sheet names into the combobox."""
        try:
            workbook = load_workbook(self.excel_file_path)
            self.sheet_combobox['values'] = workbook.sheetnames
            self.sheet_combobox.current(0)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheets: {e}")

    def on_sheet_selected(self, event):
        """Load columns when a sheet is selected."""
        self.load_columns()

    def load_columns(self):
        """Load columns from the selected sheet into the listbox."""
        self.sheet_name = self.sheet_combobox.get()
        try:
            self.df = pd.read_excel(self.excel_file_path, sheet_name=self.sheet_name)
            self.df.columns = [Utils.clean_column_name(col) for col in self.df.columns]
            self.column_listbox.delete(0, "end")
            for col in self.df.columns:
                self.column_listbox.insert("end", col)

            # After loading columns, show them in the key language combobox as options
            self.key_language_combobox['values'] = self.df.columns.tolist()
            self.key_language_combobox.current(0)  # Default the first column to be selected

            # Now, provide additional fields like database name, datatype, etc.
            self.additional_fields_frame.destroy()  # Clear previous additional fields
            self.additional_fields_frame = Frame(self.root, bg="#f5f5f5")
            self.additional_fields_frame.pack(fill="both", expand=True, pady=10)

            for i, col in enumerate(self.df.columns):
                self.create_additional_fields(i, col)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load columns: {e}")

    def create_additional_fields(self, row, col):
        """Create additional fields for database name, datatype, skip logic, and datatype value."""
        Label(self.additional_fields_frame, text=f"Column: {col}", bg="#f5f5f5").grid(row=row, column=0, sticky="w", padx=10, pady=5)

        db_name_entry = Entry(self.additional_fields_frame, width=30)
        db_name_entry.grid(row=row, column=1, padx=5, pady=5)
        db_name_entry.insert(0, col)
        self.column_mappings[f"{col}_db_name"] = db_name_entry

        datatype_combobox = Combobox(self.additional_fields_frame, width=30, values=["String", "Int", "Boolean"])
        datatype_combobox.grid(row=row, column=2, padx=5, pady=5)
        datatype_combobox.set("String")
        self.column_mappings[f"{col}_datatype"] = datatype_combobox

        skip_logic_entry = Entry(self.additional_fields_frame, width=30)
        skip_logic_entry.grid(row=row, column=3, padx=5, pady=5)
        self.column_mappings[f"{col}_skip_logic"] = skip_logic_entry

        datatype_value_entry = Entry(self.additional_fields_frame, width=30)
        datatype_value_entry.grid(row=row, column=4, padx=5, pady=5)
        self.column_mappings[f"{col}_datatype_value"] = datatype_value_entry

    def map_columns(self):
        """Map selected columns to localization keys or other field types."""
        selected_indices = self.column_listbox.curselection()
        self.selected_columns = [self.column_listbox.get(i) for i in selected_indices]

        if not self.selected_columns:
            messagebox.showerror("Error", "Please select at least one column to map.")
            return

    def generate_dart_files(self):
        """Generate Dart files based on mapped columns."""
        try:
            localization_mapping = {col: entry.get() for col, entry in self.column_mappings.items() if entry.get()}
            if not localization_mapping:
                messagebox.showerror("Error", "Please provide mappings for all selected columns.")
                return

            # Generate Dart files (this is a placeholder - add your generation logic here)
            output_folder = "dart_output"
            os.makedirs(output_folder, exist_ok=True)

            for language in self.languages:
                language_data = self.df[localization_mapping.keys()]
                dart_content = self.generate_dart_code(language, language_data)
                Utils.write_dart_file(output_folder, f"{language}.dart", dart_content)

            messagebox.showinfo("Success", "Dart files generated successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Dart files: {e}")

    def generate_dart_code(self, language, language_data):
        """Generate Dart localization code for a specific language."""
        content = f"class {language.capitalize()} {{\n"
        for key, value in language_data.items():
            content += f"  static const String {Utils.sanitize_key(key)} = '{value}';\n"
        content += "}\n"
        return content


# Start the application
if __name__ == "__main__":
    root = Tk()
    app = DartCodeGeneratorGUI(root)
    root.mainloop()
